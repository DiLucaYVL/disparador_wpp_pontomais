"""Utilitários para exportação do histórico de envios em formato Excel."""
from __future__ import annotations

from collections import defaultdict
from io import BytesIO
from typing import Dict, Iterable, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

RegistroHistorico = Dict[str, str]

HEADER_TITLES = [
    "Equipe",
    "Quantidade",
    "Nome",
    "Tipo de Relatório",
    "Motivo",
    "Quantidade Detalhe",
]

DEFAULT_TEXT = "Não informado"
DEFAULT_TYPE = "Sem tipo"
DEFAULT_REASON = "Sem motivo"
DEFAULT_PERSON = "Sem identificação"


def _normalizar(valor: str | None, padrao: str) -> str:
    if isinstance(valor, str):
        texto = valor.strip()
        return texto or padrao
    return padrao


def agrupar_envios(registros: Iterable[RegistroHistorico]) -> List[Dict[str, object]]:
    """Agrupa os registros por equipe, retornando totais e detalhes."""
    agrupado: Dict[str, Dict[str, object]] = {}

    for registro in registros:
        equipe = _normalizar(registro.get("equipe"), DEFAULT_TEXT)
        tipo = _normalizar(registro.get("tipo_relatorio"), DEFAULT_TYPE)
        motivo = _normalizar(registro.get("motivo_envio"), DEFAULT_REASON)
        pessoa = _normalizar(registro.get("pessoa"), DEFAULT_PERSON)

        equipe_map = agrupado.setdefault(
            equipe,
            {"total": 0, "detalhes": defaultdict(int)},
        )

        equipe_map["total"] = int(equipe_map["total"]) + 1
        chave = (pessoa, tipo, motivo)
        equipe_map["detalhes"][chave] += 1

    resultado: List[Dict[str, object]] = []
    for equipe in sorted(agrupado.keys(), key=lambda valor: valor.upper()):
        detalhes_map: Dict[Tuple[str, str, str], int] = agrupado[equipe]["detalhes"]
        detalhes_ordenados = sorted(
            detalhes_map.items(),
            key=lambda item: (
                item[0][0].upper(),
                item[0][1].upper(),
                item[0][2].upper(),
            ),
        )
        detalhes = [
            {
                "pessoa": pessoa,
                "tipo": tipo,
                "motivo": motivo,
                "total": quantidade,
            }
            for (pessoa, tipo, motivo), quantidade in detalhes_ordenados
        ]
        resultado.append(
            {
                "equipe": equipe,
                "total": agrupado[equipe]["total"],
                "detalhes": detalhes,
            }
        )
    return resultado


def gerar_planilha_historico(registros: Iterable[RegistroHistorico]) -> BytesIO:
    """Gera um arquivo Excel contendo o histórico formatado."""
    grupos = agrupar_envios(registros)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Histórico"
    worksheet.freeze_panes = "A2"
    outline = worksheet.sheet_properties.outlinePr
    outline.summaryBelow = False
    outline.applyStyles = True

    header_font = Font(color="FFFFFF", bold=True)
    header_fill = PatternFill("solid", fgColor="4C5BF1")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="E0E5FF")
    header_border = Border(top=thin, left=thin, right=thin, bottom=thin)

    summary_fill = PatternFill("solid", fgColor="EBF1FF")
    summary_font = Font(bold=True, color="2C3E50")
    summary_alignment = Alignment(horizontal="left")

    detail_alignment = Alignment(horizontal="left")

    worksheet.append(HEADER_TITLES)
    for col_idx, title in enumerate(HEADER_TITLES, start=1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border

    column_widths = [28, 14, 28, 22, 38, 16]
    for idx, largura in enumerate(column_widths, start=1):
        worksheet.column_dimensions[get_column_letter(idx)].width = largura

    linha_atual = 2
    if not grupos:
        celula = worksheet.cell(row=linha_atual, column=1, value="Nenhum dado encontrado para os filtros informados.")
        celula.alignment = Alignment(horizontal="left")
        celula.font = Font(color="6C757D", italic=True)
        worksheet.merge_cells(start_row=linha_atual, start_column=1, end_row=linha_atual, end_column=len(HEADER_TITLES))
    else:
        for grupo in grupos:
            summary_row = linha_atual
            worksheet.cell(row=linha_atual, column=1, value=grupo["equipe"])
            worksheet.cell(row=linha_atual, column=2, value=grupo["total"])
            for col in range(1, len(HEADER_TITLES) + 1):
                cell = worksheet.cell(row=linha_atual, column=col)
                cell.fill = summary_fill
                cell.font = summary_font
                cell.border = header_border
                if col == 2:
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = summary_alignment
            linha_atual += 1

            detalhes: List[Dict[str, object]] = grupo["detalhes"]  # type: ignore[assignment]
            for detalhe in detalhes:
                worksheet.cell(row=linha_atual, column=3, value=detalhe["pessoa"])
                worksheet.cell(row=linha_atual, column=4, value=detalhe["tipo"])
                worksheet.cell(row=linha_atual, column=5, value=detalhe["motivo"])
                quantidade_celula = worksheet.cell(row=linha_atual, column=6, value=detalhe["total"])
                quantidade_celula.alignment = Alignment(horizontal="right")

                for col in range(3, len(HEADER_TITLES) + 1):
                    cell = worksheet.cell(row=linha_atual, column=col)
                    cell.border = header_border
                    if cell.alignment.horizontal is None:
                        cell.alignment = detail_alignment

                dimensao = worksheet.row_dimensions[linha_atual]
                dimensao.outlineLevel = 1
                dimensao.hidden = True
                dimensao.collapsed = True

                linha_atual += 1

            summary_dim = worksheet.row_dimensions[summary_row]
            summary_dim.outlineLevel = 0
            summary_dim.collapsed = True

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
